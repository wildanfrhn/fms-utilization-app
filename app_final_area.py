# ============================================================
# FMS Utilization Converter - Streamlit App
# Based on original Colab notebook: Utilisasi_Fix_Juni_2026.ipynb
# Tambahan fitur:
# 1. Pilih bulan output berdasarkan Date Dept
# 2. Pilih hari libur dengan checkbox
# 3. Prioritas status jika bentrok per Lic Number + Date Dept:
#    Utilized > Standby > Permit > PM > RM
# 4. Logic hari libur dan kompensasi Utilized libur dengan Standby non-libur
# ============================================================

import calendar
from datetime import datetime
from io import BytesIO
from zipfile import ZipFile, ZIP_DEFLATED

import numpy as np
import pandas as pd
import streamlit as st

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================
# MAPPING AREA - BRANCH
# Berdasarkan file: BRANCH DAN AREA.xlsx
# ============================================================
AREA_BRANCH_MAP = {'JASUM': ['BANDA ACEH',
           'BATAM',
           'BENGKULU',
           'JAKARTA',
           'JAMBI',
           'MEDAN',
           'MEULABOH',
           'MUARA ENIM',
           'PADANG',
           'PALEMBANG',
           'PANGKAL PINANG',
           'PEKANBARU'],
 'KALIMANTAN': ['BALIKPAPAN',
                'BANJARMASIN',
                'BATUKAJANG',
                'BATULICIN',
                'MELAK',
                'PONTIANAK',
                'RANTAU',
                'SAMARINDA',
                'SANGATTA',
                'SEBAMBAN',
                'TABANG',
                'TANJUNG ADARO',
                'TANJUNG REDEP',
                'TARAKAN'],
 'SUB-EI': ['BATU HIJAU',
            'KENDARI',
            'MAKASSAR',
            'MANADO',
            'MATARAM',
            'PALU',
            'SOROAKO',
            'SORONG',
            'SURABAYA',
            'TERNATE',
            'TIMIKA',
            'WEDA BAY']}


# ============================================================
# SETUP HALAMAN
# ============================================================

st.set_page_config(
    page_title="FMS Utilization Converter",
    page_icon="🚚",
    layout="wide"
)

st.title("FMS Utilization Converter")
st.caption("Pilih area, upload Data FMS, pilih bulan output, pilih hari libur, lalu download hasil Data Utilisasi.")


# ============================================================
# FUNGSI BANTU DARI ALGORITMA AWAL
# ============================================================

def find_column(df, possible_names):
    """
    Mencari kolom berdasarkan beberapa kemungkinan nama.
    Tidak sensitif huruf besar/kecil dan spasi.
    """
    normalized_cols = {
        str(col).strip().lower().replace(" ", ""): col
        for col in df.columns
    }

    for name in possible_names:
        key = name.strip().lower().replace(" ", "")
        if key in normalized_cols:
            return normalized_cols[key]

    raise KeyError(f"Kolom tidak ditemukan. Dicari salah satu dari: {possible_names}")


def parse_date(value):
    """
    Mengubah nilai menjadi tanggal.
    Menggunakan dayfirst=True agar format Indonesia dd/mm/yyyy terbaca benar.
    """
    if pd.isna(value):
        return pd.NaT

    if isinstance(value, pd.Timestamp):
        return value

    return pd.to_datetime(value, dayfirst=True, errors="coerce")


def normalize_text(value):
    if pd.isna(value):
        return ""
    return str(value).strip()



def normalize_branch_for_area(value):
    """Normalisasi Branch untuk pencocokan area."""
    return normalize_text(value).upper()


def filter_by_selected_area(df_utilisasi, selected_area):
    """Filter data berdasarkan mapping Area - Branch."""
    allowed_branches = AREA_BRANCH_MAP.get(selected_area, [])
    df = df_utilisasi.copy()
    df["_Branch_Area_Key"] = df["Branch"].apply(normalize_branch_for_area)
    df_filtered = df[df["_Branch_Area_Key"].isin(allowed_branches)].copy()
    df_filtered = df_filtered.drop(columns=["_Branch_Area_Key"], errors="ignore").reset_index(drop=True)
    return df_filtered, allowed_branches


def normalize_deployment(value):
    """
    Normalisasi Deployment Type dari Data FMS.
    Jika Maintenance, diubah menjadi Preventive.
    """
    dep = normalize_text(value)

    mapping = {
        "Maintenance": "Preventive",
        "MAINTENANCE": "Preventive",
        "maintenance": "Preventive"
    }

    return mapping.get(dep, dep)


def get_status_column(deployment):
    """
    Menentukan status Utilized / Standby / Permit / PM / RM.
    """
    dep = normalize_text(deployment)

    if dep in ["SLA", "Adhoc/Charter", "Intercity", "Shuttle"]:
        return "Utilized"
    elif dep == "Unit Standby":
        return "Standby"
    elif dep == "Permit":
        return "Permit"
    elif dep == "Preventive":
        return "PM"
    elif dep == "Reactive":
        return "RM"
    else:
        return ""


def month_name_id(month_number):
    """
    Nama bulan bahasa Indonesia.
    """
    months = {
        1: "Januari",
        2: "Februari",
        3: "Maret",
        4: "April",
        5: "Mei",
        6: "Juni",
        7: "Juli",
        8: "Agustus",
        9: "September",
        10: "Oktober",
        11: "November",
        12: "Desember"
    }
    return months.get(int(month_number), str(month_number))


def compress_date_ranges(days):
    """
    Mengubah list tanggal menjadi format range.
    Contoh: [1, 2, 3, 5, 7, 8] menjadi "1-3, 5, 7-8"
    """
    if not days:
        return ""

    days = sorted(days)
    ranges = []

    start = days[0]
    prev = days[0]

    for day in days[1:]:
        if day == prev + 1:
            prev = day
        else:
            if start == prev:
                ranges.append(str(start))
            else:
                ranges.append(f"{start}-{prev}")
            start = day
            prev = day

    if start == prev:
        ranges.append(str(start))
    else:
        ranges.append(f"{start}-{prev}")

    return ", ".join(ranges)


def auto_format_excel(writer, sheet_name, date_columns=None):
    """
    Format tampilan Excel:
    - Header biru
    - Filter aktif
    - Freeze pane
    - Auto width
    - Date columns tetap tanggal Excel asli
    """
    if date_columns is None:
        date_columns = []

    ws = writer.sheets[sheet_name]

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    header_fill = PatternFill("solid", fgColor="0070C0")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    for date_col_name in date_columns:
        date_col_idx = None
        for cell in ws[1]:
            if cell.value == date_col_name:
                date_col_idx = cell.column
                break

        if date_col_idx:
            col_letter = get_column_letter(date_col_idx)
            for row in range(2, ws.max_row + 1):
                ws[f"{col_letter}{row}"].number_format = "DD/MM/YYYY"

    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 40)


# ============================================================
# FUNGSI PEMROSESAN DATA
# ============================================================

def read_uploaded_file(uploaded_file):
    """Membaca file Excel atau CSV dari Streamlit uploader."""
    filename = uploaded_file.name.lower()

    if filename.endswith(".xlsx"):
        df = pd.read_excel(uploaded_file, engine="openpyxl")
    elif filename.endswith(".xls"):
        df = pd.read_excel(uploaded_file, engine="xlrd")
    elif filename.endswith(".csv"):
        df = pd.read_csv(uploaded_file)
    else:
        raise ValueError("Format file tidak didukung. Gunakan .xlsx, .xls, atau .csv")

    df.columns = df.columns.astype(str).str.strip()
    return df


def build_daily_utilization(df_fms):
    """
    Algoritma awal: ubah Data FMS menjadi Data Utilisasi harian.
    """
    col_branch = find_column(df_fms, ["Branch"])
    col_lic = find_column(df_fms, ["Lic Number", "License Number", "Plat Nomor", "Plat Nomer"])
    col_deployment = find_column(df_fms, ["Deployment Type", "Deployment"])
    col_depart_date = find_column(df_fms, ["Depart Date", "Date Dept", "Departure Date"])
    col_arrive_date = find_column(df_fms, ["Arrive Date", "Arrival Date"])

    df_fms = df_fms.copy()
    df_fms[col_depart_date] = df_fms[col_depart_date].apply(parse_date)
    df_fms[col_arrive_date] = df_fms[col_arrive_date].apply(parse_date)

    df_fms = df_fms.dropna(subset=[col_depart_date, col_arrive_date]).copy()

    df_fms[col_branch] = df_fms[col_branch].apply(normalize_text)
    df_fms[col_lic] = df_fms[col_lic].apply(normalize_text)
    df_fms[col_deployment] = df_fms[col_deployment].apply(normalize_deployment)

    rows = []

    for _, row in df_fms.iterrows():
        depart_date = row[col_depart_date]
        arrive_date = row[col_arrive_date]

        if pd.isna(depart_date) or pd.isna(arrive_date):
            continue

        if arrive_date < depart_date:
            continue

        branch = row[col_branch]
        lic_number = row[col_lic]
        deployment = row[col_deployment]
        date_range = pd.date_range(start=depart_date, end=arrive_date, freq="D")

        for date_dept in date_range:
            status_col = get_status_column(deployment)

            utilized = ""
            standby = ""
            permit = ""
            pm = ""
            rm = ""

            if status_col == "Utilized":
                utilized = 1
            elif status_col == "Standby":
                standby = 1
            elif status_col == "Permit":
                permit = 1
            elif status_col == "PM":
                pm = 1
            elif status_col == "RM":
                rm = 1

            day_name = date_dept.day_name()
            days_status = "Holiday" if day_name == "Sunday" else "Working Day"
            remarks = f"{deployment}{status_col}"
            code = f"{date_dept.strftime('%d/%m/%Y')}{branch}{lic_number}{status_col}"
            cek_double = f"{date_dept.strftime('%d/%m/%Y')}{lic_number}"

            rows.append({
                "Date Dept": date_dept,
                "Branch": branch,
                "Lic Number": lic_number,
                "Utilized": utilized,
                "Standby": standby,
                "Permit": permit,
                "PM": pm,
                "RM": rm,
                "Deployment": deployment,
                "Month": date_dept.month,
                "Days": days_status,
                "Remarks": remarks,
                "Remarks QA": "",
                "Code": code,
                "T/F": "",
                "Valuation": "",
                "Cek Double": cek_double,
                "Day": day_name,
                "Flag Holiday": ""
            })

    df_utilisasi = pd.DataFrame(rows)

    if df_utilisasi.empty:
        raise ValueError("Data Utilisasi kosong. Cek kembali kolom tanggal Depart Date dan Arrive Date pada Data FMS.")

    return df_fms, df_utilisasi


def remove_exact_duplicates_original(df_utilisasi):
    """Algoritma awal: hapus duplikat berdasarkan semua kolom kecuali Code dan T/F."""
    df_utilisasi = df_utilisasi.sort_values(
        by=["Date Dept", "Branch", "Lic Number", "Deployment"],
        ascending=True
    ).reset_index(drop=True)

    duplicate_check_cols = [
        col for col in df_utilisasi.columns
        if col not in ["Code", "T/F"]
    ]

    return df_utilisasi.drop_duplicates(
        subset=duplicate_check_cols,
        keep="first"
    ).reset_index(drop=True)


def apply_status_priority(df_utilisasi):
    """
    Tambahan revisi:
    Untuk Lic Number + Date Dept yang sama, hanya boleh ada 1 status.
    Prioritas: Utilized > Standby > Permit > PM > RM.
    """
    df = df_utilisasi.copy()

    priority_map = {
        "Utilized": 1,
        "Standby": 2,
        "Permit": 3,
        "PM": 4,
        "RM": 5,
        "": 99
    }

    def row_status(row):
        for status in ["Utilized", "Standby", "Permit", "PM", "RM"]:
            if row.get(status, "") == 1:
                return status
        return ""

    df["_Status"] = df.apply(row_status, axis=1)
    df["_Priority"] = df["_Status"].map(priority_map).fillna(99)

    df = df.sort_values(
        by=["Date Dept", "Lic Number", "_Priority", "Branch", "Deployment"],
        ascending=True
    ).reset_index(drop=True)

    df = df.drop_duplicates(
        subset=["Date Dept", "Lic Number"],
        keep="first"
    ).reset_index(drop=True)

    df = df.drop(columns=["_Status", "_Priority"])
    return df


def recalculate_tf(df_utilisasi):
    """Algoritma awal: TRUE jika Code baris ini sama dengan Code satu baris di bawahnya."""
    df = df_utilisasi.sort_values(
        by=["Date Dept", "Branch", "Lic Number", "Deployment"],
        ascending=True
    ).reset_index(drop=True)

    df["T/F"] = np.where(
        df["Code"] == df["Code"].shift(-1),
        True,
        False
    )

    return df


def month_options_from_data(df_utilisasi):
    """Membuat pilihan bulan berdasarkan Date Dept yang tersedia."""
    temp = df_utilisasi[["Date Dept"]].copy()
    temp["Year"] = temp["Date Dept"].dt.year
    temp["Month_Num"] = temp["Date Dept"].dt.month
    temp = temp.drop_duplicates(subset=["Year", "Month_Num"]).sort_values(["Year", "Month_Num"])

    options = []
    for _, row in temp.iterrows():
        year = int(row["Year"])
        month_num = int(row["Month_Num"])
        label = f"{month_name_id(month_num)} {year}"
        value = f"{year}-{month_num:02d}"
        options.append((label, value, year, month_num))

    return options


def apply_selected_month_filter(df_utilisasi, year, month_num):
    """Filter output berdasarkan bulan dari Date Dept."""
    return df_utilisasi[
        (df_utilisasi["Date Dept"].dt.year == int(year)) &
        (df_utilisasi["Date Dept"].dt.month == int(month_num))
    ].copy().reset_index(drop=True)


def apply_holiday_logic(df_month, selected_holiday_days):
    """
    Tambahan revisi hari libur:
    - Hari libur dipilih user lewat checkbox.
    - Target Days = jumlah hari dalam bulan - jumlah hari libur.
    - Jika tanggal libur berstatus Standby/Permit/PM/RM, angka 1 menjadi 0.
    - Jika tanggal libur berstatus Utilized, cari kompensasi dari Standby di hari non-libur
      pada Lic Number yang sama.
    - Jika Standby non-libur tidak cukup, sisa kompensasi diambil dari Utilized hari libur itu sendiri.
    """
    df = df_month.copy().reset_index(drop=True)
    selected_holiday_days = set(int(day) for day in selected_holiday_days)

    if df.empty:
        return df, pd.DataFrame()

    df["_Day_Num"] = df["Date Dept"].dt.day
    df["_Is_Selected_Holiday"] = df["_Day_Num"].isin(selected_holiday_days)

    # Update kolom Days dan Flag Holiday sesuai pilihan user.
    df.loc[df["_Is_Selected_Holiday"], "Days"] = "Holiday"
    df.loc[df["_Is_Selected_Holiday"], "Flag Holiday"] = "Selected Holiday"

    adjustment_rows = []

    for lic_number in sorted(df["Lic Number"].dropna().unique()):
        lic_mask = df["Lic Number"] == lic_number
        holiday_mask = lic_mask & df["_Is_Selected_Holiday"]
        non_holiday_mask = lic_mask & (~df["_Is_Selected_Holiday"])

        # 1) Hari libur dengan status non-utilized langsung jadi 0.
        for status in ["Standby", "Permit", "PM", "RM"]:
            idx_to_zero = df.index[holiday_mask & (df[status] == 1)].tolist()
            if idx_to_zero:
                df.loc[idx_to_zero, status] = 0
                for idx in idx_to_zero:
                    adjustment_rows.append({
                        "Lic Number": lic_number,
                        "Date Dept": df.at[idx, "Date Dept"],
                        "Adjustment Type": f"Holiday {status} set to 0",
                        "Status Adjusted": status,
                        "Reason": "Tanggal dipilih sebagai hari libur"
                    })

        # 2) Utilized pada hari libur dikompensasi dengan Standby di hari non-libur.
        holiday_utilized_idx = df.index[holiday_mask & (df["Utilized"] == 1)].tolist()
        needed_compensation = len(holiday_utilized_idx)

        if needed_compensation == 0:
            continue

        standby_non_holiday_idx = df.index[non_holiday_mask & (df["Standby"] == 1)].tolist()
        standby_non_holiday_idx = sorted(standby_non_holiday_idx, key=lambda idx: df.at[idx, "Date Dept"])

        use_standby_idx = standby_non_holiday_idx[:needed_compensation]

        if use_standby_idx:
            df.loc[use_standby_idx, "Standby"] = 0
            for idx in use_standby_idx:
                adjustment_rows.append({
                    "Lic Number": lic_number,
                    "Date Dept": df.at[idx, "Date Dept"],
                    "Adjustment Type": "Standby non-holiday set to 0",
                    "Status Adjusted": "Standby",
                    "Reason": "Kompensasi untuk Utilized yang jatuh pada hari libur"
                })

        remaining = needed_compensation - len(use_standby_idx)

        if remaining > 0:
            # Jika Standby tidak cukup, sisa dikurangi dari Utilized hari libur itu sendiri.
            holiday_utilized_idx = sorted(holiday_utilized_idx, key=lambda idx: df.at[idx, "Date Dept"])
            use_utilized_idx = holiday_utilized_idx[:remaining]
            df.loc[use_utilized_idx, "Utilized"] = 0

            for idx in use_utilized_idx:
                adjustment_rows.append({
                    "Lic Number": lic_number,
                    "Date Dept": df.at[idx, "Date Dept"],
                    "Adjustment Type": "Holiday Utilized set to 0",
                    "Status Adjusted": "Utilized",
                    "Reason": "Standby non-libur tidak cukup untuk kompensasi"
                })

    df = df.drop(columns=["_Day_Num", "_Is_Selected_Holiday"])

    df_adjustment = pd.DataFrame(
        adjustment_rows,
        columns=["Lic Number", "Date Dept", "Adjustment Type", "Status Adjusted", "Reason"]
    )

    return df, df_adjustment


def build_keterangan_kosong(df_source):
    """
    Algoritma awal file keterangan tanggal kosong.
    Untuk setiap Lic Number dan bulan yang muncul, dicek tanggal mana yang belum ada.
    """
    gap_rows = []

    if df_source.empty:
        return pd.DataFrame(columns=["Lic Number", "Keterangan"])

    df_gap_source = df_source.copy()
    df_gap_source["Year"] = df_gap_source["Date Dept"].dt.year
    df_gap_source["Month_Num"] = df_gap_source["Date Dept"].dt.month
    df_gap_source["Day_Num"] = df_gap_source["Date Dept"].dt.day

    for lic_number in sorted(df_gap_source["Lic Number"].dropna().unique()):
        df_lic = df_gap_source[df_gap_source["Lic Number"] == lic_number]
        keterangan_list = []

        month_year_pairs = (
            df_lic[["Year", "Month_Num"]]
            .drop_duplicates()
            .sort_values(["Year", "Month_Num"])
            .values
        )

        for year, month_num in month_year_pairs:
            df_month = df_lic[
                (df_lic["Year"] == year) &
                (df_lic["Month_Num"] == month_num)
            ]

            existing_days = set(df_month["Day_Num"].dropna().astype(int).tolist())
            last_day = calendar.monthrange(int(year), int(month_num))[1]
            full_days = set(range(1, last_day + 1))
            missing_days = sorted(list(full_days - existing_days))

            if missing_days:
                range_text = compress_date_ranges(missing_days)
                bulan_text = month_name_id(int(month_num))
                keterangan_list.append(
                    f"Ops belum menginput tgl {range_text} {bulan_text} di FMS"
                )

        if keterangan_list:
            gap_rows.append({
                "Lic Number": lic_number,
                "Keterangan": "; ".join(keterangan_list)
            })

    return pd.DataFrame(gap_rows, columns=["Lic Number", "Keterangan"])


def build_summary(df_output, target_days):
    """Ringkasan total status per Lic Number untuk validasi Target Days."""
    status_cols = ["Utilized", "Standby", "Permit", "PM", "RM"]
    if df_output.empty:
        return pd.DataFrame(columns=["Lic Number"] + status_cols + ["Total Status", "Target Days", "Check"])

    df_calc = df_output.copy()
    for col in status_cols:
        df_calc[col] = pd.to_numeric(df_calc[col], errors="coerce").fillna(0).astype(int)

    summary = (
        df_calc.groupby("Lic Number", as_index=False)[status_cols]
        .sum()
        .sort_values("Lic Number")
    )
    summary["Total Status"] = summary[status_cols].sum(axis=1)
    summary["Target Days"] = int(target_days)
    summary["Check"] = np.select(
        [
            summary["Total Status"] == int(target_days),
            summary["Total Status"] < int(target_days),
            summary["Total Status"] > int(target_days),
        ],
        ["OK", "Kurang dari Target Days", "Melebihi Target Days"],
        default="Tidak Sesuai Target Days"
    )

    return summary


def finalize_columns(df_utilisasi):
    """Susun kolom final seperti algoritma awal."""
    final_columns = [
        "Date Dept",
        "Branch",
        "Lic Number",
        "Utilized",
        "Standby",
        "Permit",
        "PM",
        "RM",
        "Deployment",
        "Month",
        "Days",
        "Remarks",
        "Remarks QA",
        "Code",
        "T/F",
        "Valuation",
        "Cek Double",
        "Day",
        "Flag Holiday"
    ]

    return df_utilisasi[final_columns].copy()


def dataframe_to_excel_bytes(sheets):
    """Membuat file Excel dalam bentuk bytes untuk download Streamlit."""
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl", datetime_format="DD/MM/YYYY") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            date_cols = [col for col in df.columns if col == "Date Dept"]
            auto_format_excel(writer, sheet_name=sheet_name, date_columns=date_cols)

    output.seek(0)
    return output.getvalue()


def create_zip_bytes(files_dict):
    """Membuat ZIP berisi beberapa file output."""
    zip_buffer = BytesIO()
    with ZipFile(zip_buffer, "w", ZIP_DEFLATED) as zf:
        for filename, file_bytes in files_dict.items():
            zf.writestr(filename, file_bytes)
    zip_buffer.seek(0)
    return zip_buffer.getvalue()


# ============================================================
# UI STREAMLIT
# ============================================================

with st.sidebar:
    st.header("Input")

    selected_area = st.selectbox(
        "Pilih Area",
        options=["KALIMANTAN", "SUB-EI", "JASUM"],
        help="Output hanya akan menampilkan Branch yang masuk ke area terpilih."
    )

    uploaded_file = st.file_uploader(
        "Upload Data FMS",
        type=["xlsx", "xls", "csv"],
        help="Upload file Data FMS mentah dalam format Excel atau CSV."
    )

if uploaded_file is None:
    st.info("Silakan upload file Data FMS terlebih dahulu.")
    st.stop()

try:
    df_fms_raw = read_uploaded_file(uploaded_file)

    st.subheader("Preview Data FMS")
    st.write(f"Jumlah baris terbaca: **{len(df_fms_raw):,}**")
    st.dataframe(df_fms_raw.head(20), use_container_width=True)

    df_fms_clean, df_daily = build_daily_utilization(df_fms_raw)
    df_daily = remove_exact_duplicates_original(df_daily)
    df_daily = apply_status_priority(df_daily)
    df_daily = recalculate_tf(df_daily)

    df_daily, allowed_branches = filter_by_selected_area(df_daily, selected_area)

    st.subheader("Area Filter")
    st.write(f"Area dipilih: **{selected_area}**")
    st.write(f"Branch dalam area ini: {', '.join(allowed_branches)}")
    st.write(f"Baris setelah filter area: **{len(df_daily):,}**")

    if df_daily.empty:
        st.error("Tidak ada data yang cocok dengan area yang dipilih. Cek kolom Branch pada file FMS.")
        st.stop()

    options = month_options_from_data(df_daily)
    if not options:
        st.error("Tidak ada bulan yang terbaca dari Date Dept.")
        st.stop()

    st.subheader("Pilih Bulan Output")
    option_labels = [opt[0] for opt in options]
    selected_label = st.selectbox("Bulan output berdasarkan Date Dept", option_labels)
    selected_option = next(opt for opt in options if opt[0] == selected_label)
    _, _, selected_year, selected_month = selected_option

    last_day = calendar.monthrange(int(selected_year), int(selected_month))[1]
    all_days = list(range(1, last_day + 1))

    st.subheader("Pilih Tanggal Hari Libur")
    st.caption("Tanggal yang dicentang akan dianggap sebagai hari libur pada bulan output yang dipilih.")

    default_sundays = [
        day for day in all_days
        if datetime(int(selected_year), int(selected_month), int(day)).weekday() == 6
    ]

    with st.expander("Centang tanggal hari libur", expanded=True):
        selected_holiday_days = st.multiselect(
            "Tanggal libur",
            options=all_days,
            default=default_sundays,
            format_func=lambda day: f"{day} {month_name_id(selected_month)} {selected_year}"
        )

    target_days = last_day - len(selected_holiday_days)

    col1, col2, col3 = st.columns(3)
    col1.metric("Jumlah Hari Bulan", last_day)
    col2.metric("Hari Libur Dipilih", len(selected_holiday_days))
    col3.metric("Target Days", target_days)

    process_clicked = st.button("Process Data", type="primary")

    if not process_clicked:
        st.stop()

    df_month = apply_selected_month_filter(df_daily, selected_year, selected_month)
    df_month_adjusted, df_adjustment = apply_holiday_logic(df_month, selected_holiday_days)
    df_month_adjusted = recalculate_tf(df_month_adjusted)
    df_utilisasi_output = finalize_columns(df_month_adjusted)
    df_keterangan = build_keterangan_kosong(df_month_adjusted)
    df_summary = build_summary(df_utilisasi_output, target_days)

    st.success("Proses selesai.")

    st.subheader("Ringkasan Output")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Baris Data FMS", f"{len(df_fms_clean):,}")
    c2.metric("Baris Output", f"{len(df_utilisasi_output):,}")
    c3.metric("Unique Lic Number", f"{df_utilisasi_output['Lic Number'].nunique():,}")
    c4.metric("Lic Number dengan Tgl Kosong", f"{len(df_keterangan):,}")

    not_ok_count = int((df_summary["Check"] != "OK").sum()) if not df_summary.empty else 0
    if not_ok_count > 0:
        st.warning(f"Ada {not_ok_count} Lic Number yang total statusnya belum sesuai Target Days. Cek sheet Summary.")
    else:
        st.success("Semua Lic Number sudah sesuai Target Days.")

    tab1, tab2, tab3, tab4 = st.tabs(["Data Utilisasi", "Summary", "Keterangan Kosong", "Adjustment Log"])

    with tab1:
        st.dataframe(df_utilisasi_output.head(500), use_container_width=True)
        st.caption("Preview maksimal 500 baris pertama.")

    with tab2:
        st.dataframe(df_summary, use_container_width=True)

    with tab3:
        st.dataframe(df_keterangan, use_container_width=True)

    with tab4:
        st.dataframe(df_adjustment, use_container_width=True)

    utilisasi_bytes = dataframe_to_excel_bytes({
        "Data Utilisasi": df_utilisasi_output,
        "Summary": df_summary,
        "Adjustment Log": df_adjustment
    })

    keterangan_bytes = dataframe_to_excel_bytes({
        "Keterangan": df_keterangan
    })

    zip_bytes = create_zip_bytes({
        "Data_Utilisasi_Output.xlsx": utilisasi_bytes,
        "Data_Keterangan_Kosong_FMS.xlsx": keterangan_bytes
    })

    st.subheader("Download Output")
    d1, d2, d3 = st.columns(3)

    with d1:
        st.download_button(
            label="Download Data Utilisasi",
            data=utilisasi_bytes,
            file_name="Data_Utilisasi_Output.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    with d2:
        st.download_button(
            label="Download Keterangan Kosong",
            data=keterangan_bytes,
            file_name="Data_Keterangan_Kosong_FMS.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    with d3:
        st.download_button(
            label="Download Semua Output ZIP",
            data=zip_bytes,
            file_name="Output_FMS_Utilization.zip",
            mime="application/zip"
        )

except Exception as e:
    st.error("Terjadi error saat memproses data.")
    st.exception(e)
